import openpyxl as op
from xlutils.copy import copy
import pandas as pd
from common import ReadYaml, Path
import xlrd

way = Path.project_path()

class ExcelWork:

    def __init__(self, file_path, sheet_name):
        self.file_path = file_path
        self.wb = op.load_workbook(file_path)
        self.max_col = self.get_max_col(sheet_name)
        self.sheet_data = self.wb[sheet_name]

    def get_max_col(self, sheet_name):
        self.sheet_data = self.wb[sheet_name]
        max_col = self.sheet_data.max_column
        return max_col + 1

    def write_data(self, x, data):
        if x == None:
            pass
        else:
            # max_col = sheet_data.max_column
            # max_row = sheet_data.max_row
            self.sheet_data.cell(1, self.max_col, '结果')
            self.sheet_data.cell(x, self.max_col, data)  # x 行
            # 保存
            self.wb.save(self.file_path)

    def writeExcel(self, data):
        for i in data:
            eval("self.sheet_data.cell{}".format(i))  # x 行, y列 (x,y,值')
        # 保存
        self.wb.save(self.file_path)

    def pd_read(self):
        data = pd.read_excel(self.file_path)
        # 替换nan
        data.fillna('null', inplace=True)
        # df = pd.DataFrame(data)
        # df.to_excel()
        header = data.columns.to_list()
        row1 = data.values[0]
        row2 = data.values[1]
        row1 = dict(zip(header, row1))
        row2 = dict(zip(header, row2))
        print(header)
        print(row1)
        print(row2)
        dataSalary = {}  # 第一行与第二行不同的值
        for i in header:
            if row1[i] == row2[i]:
                continue
            else:
                dataSalary[i] = row2[i]  # 取出第二行计算后的，不同的值
        print(dataSalary)
        # print(i, row1[i])
        # self.calculation(dataSalary)

    # 计算
    def calculation(self, dataSalary):
        value = dataSalary['应发金额']
        exceptValue = 1
        assert exceptValue == value, '应发金额计算错误！'

if __name__ == '__main__':
    file_path = r'{}/enclosure/薪酬模板.xlsx'.format(way)
    excel = ExcelWork(file_path, 'sheet1')
    excel.pd_read()
    # data = [(2, 1, '姓名'), (2, 2, '工号'), (2, 3, '人员单位名称'), (2, 4, '人员机构名称')]
    # excel.writeExcel(data)
    # excel.write_data(2, 'PASS')
